#!/usr/bin/python3.5.2
# -*- coding: utf-8 -*-
"""
@Time    : 2018/11/15 10:24
@Author  : TX
@File    : gpu_info_2.py
@Software: PyCharm
"""
import time

import xlsxwriter
from py3nvml import nvidia_smi
from openpyxl import load_workbook, Workbook

"""每秒获取一次GPU的信息并写入表格中"""


class GPUHandler(object):
    """获取GPU信息，单例"""
    __instance = None
    __first_init = False

    def __new__(cls):
        if not cls.__instance:
            cls.__instance = object.__new__(cls)
        return cls.__instance

    def __init__(self):
        if not GPUHandler.__first_init:
            self.my_nvidia_smi = nvidia_smi
            self.my_nvidia_smi.nvmlInit()  # 显卡初始化
            GPUHandler.__first_init = True

    def get_gpu_info(self):
        gpu_count = self.my_nvidia_smi.nvmlDeviceGetCount()  # 获取显卡总数
        # [{'gpu_num':0, 'gpu_name': 'xxx', 'mem_total':'4096MB', 'mem_used':'322MB','mem_percent': '20%'},{}....]
        gpu_info_list = []
        for i in range(gpu_count):
            handle = self.my_nvidia_smi.nvmlDeviceGetHandleByIndex(i)  # 创建操作某个GPU的对象
            memory_info = self.my_nvidia_smi.nvmlDeviceGetMemoryInfo(handle)
            gpu_num = i
            gpu_name = self.my_nvidia_smi.nvmlDeviceGetName(handle)
            mem_total = int(memory_info.total / 1024 / 1024)
            mem_used = int(memory_info.used / 1024 / 1024)
            mem_percent = memory_info.used / memory_info.total * 100
            gpu_dict = dict(gpu_num=str(gpu_num), gpu_name=gpu_name, mem_total=str(mem_total) + 'M',
                            mem_used=str(mem_used) + 'M', mem_percent=str('%.2f' % mem_percent) + '%')
            gpu_info_list.append(gpu_dict)
        return gpu_info_list


if __name__ == "__main__":
    now_time = time.strftime('%Y-%m-%d %H:%M:%S')
    excel_name = now_time.replace(' ', '').replace(':', '').replace('-', '') + '.xlsx'
    first_workbook = xlsxwriter.Workbook(excel_name)  # 可以生成.xls文件但是会报错
    first_worksheet = first_workbook.add_worksheet('Sheet1')  # 工作页
    bold = first_workbook.add_format({'bold': False})
    head = ['时间', 'GPU名称', '显存占用(M)', 'GPU利用率']
    first_worksheet.write_row('A1', head)
    first_workbook.close()

    # wb = Workbook(excel_name)
    # wb.create_sheet(title='sheet1')
    # ws = wb.active
    # head = ['时间', 'GPU名称', '显存占用(M)', 'GPU利用率']
    # ws.append(head)

    wb = load_workbook(excel_name)
    ws = wb.worksheets[0]
    ws.append(['1', '2', '3', '4'])
    wb.save(excel_name)
    ws.append([2, 2, 2, 2])
    wb.save(excel_name)

    print('ok')
    # gpu_handler = GPUHandler()
    # while True:
    #     # 获取gpu信息，构造数据写入表格中
    #     try:
    #         workbook, worksheet, sheet_row_num = write_excel_obj(excel_name)
    #         count = sheet_row_num
    #         for x in range(5):
    #             all_gpu_info = gpu_handler.get_gpu_info()
    #             if all_gpu_info:
    #                 for gpu_info_dict in all_gpu_info:
    #                     # info = []
    #                     current_time = time.strftime('%Y-%m-%d %H:%M:%S')
    #                     gpu_number = gpu_info_dict['gpu_num']
    #                     gpu_name = gpu_info_dict['gpu_name']
    #                     name = gpu_number + '-' + gpu_name
    #                     gpu_mem_used = gpu_info_dict['mem_used']
    #                     gpu_mem_total = gpu_info_dict['mem_total']
    #                     used = gpu_mem_used + '/' + gpu_mem_total
    #                     gpu_mem_percent = gpu_info_dict['mem_percent']
    #                     info_list = [current_time, name, used, gpu_mem_percent]
    #                     next_row = 'A' + str(count)
    #                     worksheet.write_row(next_row, info_list)
    #                     count += 1
    #             print(all_gpu_info)
    #             time.sleep(1)
    #         workbook.close()
    #     except Exception as e:
    #         print(e)
